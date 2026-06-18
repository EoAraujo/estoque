"""
Views do módulo de relatórios (Fase 4).

Exporta dados em Excel (.xlsx) e PDF, com filtros opcionais.
"""
from datetime import timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.timezone import localdate
from django.views.decorators.http import require_GET

from accounts.models import UserProfile
from audit.models import LogAuditoria
from core.models import Categoria, Fornecedor, Produto
from stock.models import Lote, Movimento
from stock.services import valor_medio_unitario, valor_total_estoque_produto

from .utils import apply_borders, fmt_date, fmt_decimal


def _get_periodo(request):
    """Resolve o período (dias) de análise: ?periodo= ou UserProfile.periodo_padrao."""
    raw = request.GET.get("periodo")
    if raw and raw.isdigit():
        dias = int(raw)
        if dias in dict(UserProfile.PERIODO_CHOICES):
            return dias
    profile = getattr(request.user, "profile", None)
    if profile:
        return profile.periodo_padrao
    return UserProfile.PERIODO_PADRAO


def _xlsx_response(ws, filename):
    """Empacota uma Worksheet em HttpResponse com content-type xlsx."""
    from openpyxl.workbook import Workbook
    from openpyxl.utils import get_column_letter

    if not isinstance(ws, Workbook):
        wb = Workbook()
        wb.active.title = ws.title
        wb.active._cells = ws._cells
        wb.active.column_dimensions = ws.column_dimensions
    else:
        wb = ws

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ============================================================================
# Export: Movimentações (Excel)
# ============================================================================
@login_required
@require_GET
def exportar_movimentacoes_xlsx(request):
    """Exporta o histórico de movimentações para Excel.

    Filtros via querystring: ?periodo=30&tipo=ENTRADA|SAIDA&produto=<id>
                            &categoria=<id>&inicio=YYYY-MM-DD&fim=YYYY-MM-DD
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    qs = Movimento.objects.select_related(
        "produto", "produto__categoria", "lote", "created_by", "updated_by",
    ).order_by("-data_movimento")

    # Filtros
    tipo = request.GET.get("tipo")
    if tipo in ("ENTRADA", "SAIDA", "AJUSTE"):
        qs = qs.filter(tipo=tipo)

    produto_id = request.GET.get("produto")
    if produto_id and produto_id.isdigit():
        qs = qs.filter(produto_id=int(produto_id))

    categoria_id = request.GET.get("categoria")
    if categoria_id and categoria_id.isdigit():
        qs = qs.filter(produto__categoria_id=int(categoria_id))

    inicio = request.GET.get("inicio")
    fim = request.GET.get("fim")
    if inicio:
        try:
            qs = qs.filter(data_movimento__date__gte=inicio)
        except Exception:
            pass
    if fim:
        try:
            qs = qs.filter(data_movimento__date__lte=fim)
        except Exception:
            pass
    else:
        dias = _get_periodo(request)
        desde = timezone.now() - timedelta(days=dias)
        qs = qs.filter(data_movimento__gte=desde)

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentações"

    # Cabeçalho
    headers = [
        "Data/Hora", "Tipo", "Produto", "Código", "Categoria",
        "Lote", "Qtd", "Unidade", "Valor Unit.", "Valor Total",
        "Motivo", "Cancelado", "Usuário",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="065F46")  # emerald-800
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Linhas
    for m in qs:
        ws.append([
            fmt_date(m.data_movimento),
            m.get_tipo_display(),
            m.produto.nome,
            m.produto.codigo_interno,
            m.produto.categoria.nome if m.produto.categoria_id else "",
            m.lote.numero_lote if m.lote_id else "",
            fmt_decimal(m.quantidade, 3),
            m.produto.unidade_medida,
            fmt_decimal(m.valor_unitario, 2),
            fmt_decimal(m.valor_total, 2),
            m.motivo or "",
            "Sim" if m.cancelado else "Não",
            m.created_by.get_full_name() if m.created_by else (
                m.created_by.username if m.created_by else ""
            ),
        ])

    # Totais
    total_qtd_entrada = sum(
        (m.quantidade for m in qs if m.tipo == "ENTRADA" and not m.cancelado),
        Decimal("0"),
    )
    total_qtd_saida = sum(
        (m.quantidade for m in qs if m.tipo == "SAIDA" and not m.cancelado),
        Decimal("0"),
    )
    total_valor = sum(
        (m.valor_total for m in qs if m.tipo == "ENTRADA" and not m.cancelado),
        Decimal("0"),
    )
    ws.append([])
    ws.append(["", "Totais (não cancelados):", "", "", "", "",
               f"E:{fmt_decimal(total_qtd_entrada,3)} S:{fmt_decimal(total_qtd_saida,3)}",
               "", "", fmt_decimal(total_valor, 2), "", "", ""])

    # Larguras
    widths = [18, 12, 30, 14, 16, 14, 10, 8, 12, 14, 30, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_row = ws.max_row
    last_col = get_column_letter(len(headers))
    apply_borders(ws, f"A1:{last_col}{last_row}")
    ws.freeze_panes = "A2"

    filename = f"movimentacoes_{localdate().isoformat()}.xlsx"
    return _xlsx_response(wb, filename)


# ============================================================================
# Export: Posição de Estoque (Excel)
# ============================================================================
@login_required
@require_GET
def exportar_posicao_estoque_xlsx(request):
    """Exporta posição de estoque (produto + lotes) para Excel.

    Filtros: ?categoria=<id>&fornecedor=<id>&apenas_com_estoque=1
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    categoria_id = request.GET.get("categoria")
    fornecedor_id = request.GET.get("fornecedor")
    apenas_com_estoque = request.GET.get("apenas_com_estoque") == "1"

    produtos = Produto.objects.select_related(
        "categoria", "fornecedor_principal",
    ).filter(ativo=True).order_by("nome")

    if categoria_id and categoria_id.isdigit():
        produtos = produtos.filter(categoria_id=int(categoria_id))
    if fornecedor_id and fornecedor_id.isdigit():
        produtos = produtos.filter(fornecedor_principal_id=int(fornecedor_id))

    wb = Workbook()
    ws = wb.active
    ws.title = "Posição de Estoque"

    headers = [
        "Código", "Produto", "Categoria", "Unidade", "Estoque Atual",
        "Estoque Mínimo", "Estoque Ideal", "Valor Unit. Médio",
        "Valor Total", "Localização", "Controla Validade",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="065F46")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    total_geral = Decimal("0")
    for p in produtos:
        qtd = p.quantidade_atual
        if apenas_com_estoque and qtd <= 0:
            continue
        vlr_unit = valor_medio_unitario(p) or Decimal("0")
        vlr_total = valor_total_estoque_produto(p)
        total_geral += vlr_total
        ws.append([
            p.codigo_interno,
            p.nome,
            p.categoria.nome if p.categoria_id else "",
            p.unidade_medida,
            fmt_decimal(qtd, 3),
            fmt_decimal(p.estoque_minimo, 3),
            fmt_decimal(p.estoque_ideal, 3),
            fmt_decimal(vlr_unit, 2),
            fmt_decimal(vlr_total, 2),
            p.localizacao or "",
            "Sim" if p.controla_validade else "Não",
        ])

    # Aba de lotes
    ws2 = wb.create_sheet("Lotes")
    headers2 = [
        "Produto", "Código", "Lote", "Qtd. Atual", "Data Entrada",
        "Data Validade", "Dias p/ Vencer", "Status", "Valor Unit.",
    ]
    ws2.append(headers2)
    for col_idx, _ in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    hoje = localdate()
    for p in produtos:
        if apenas_com_estoque and p.quantidade_atual <= 0:
            continue
        vlr_unit_p = valor_medio_unitario(p) or Decimal("0")
        for lote in Lote.objects.filter(produto=p, ativo=True).order_by("data_validade"):
            dias = lote.dias_para_vencer
            if lote.vencido:
                status = "VENCIDO"
            elif dias is not None and dias <= 3:
                status = "Crítico"
            elif dias is not None and dias <= 7:
                status = "Alerta"
            elif dias is not None and dias <= 30:
                status = "Atenção"
            else:
                status = "OK"
            ws2.append([
                p.nome,
                p.codigo_interno,
                lote.numero_lote,
                fmt_decimal(lote.quantidade_atual, 3),
                fmt_date(lote.data_entrada),
                fmt_date(lote.data_validade),
                dias if dias is not None else "",
                status,
                fmt_decimal(vlr_unit_p, 2),
            ])

    # Rodapé com total
    ws.append([])
    ws.append(["", "TOTAL GERAL (R$):", "", "", "", "", "", "", fmt_decimal(total_geral, 2), "", ""])

    # Larguras
    widths = [14, 30, 16, 8, 14, 14, 14, 14, 14, 18, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    widths2 = [30, 14, 14, 12, 14, 14, 14, 12, 12]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    apply_borders(ws, f"A1:K{ws.max_row}")
    apply_borders(ws2, f"A1:I{ws2.max_row}")
    ws.freeze_panes = "A2"
    ws2.freeze_panes = "A2"

    filename = f"posicao_estoque_{localdate().isoformat()}.xlsx"
    return _xlsx_response(wb, filename)


# ============================================================================
# Export: Auditoria (Excel)
# ============================================================================
@login_required
@require_GET
def exportar_auditoria_xlsx(request):
    """Exporta logs de auditoria para Excel.

    Filtros: ?acao=CREATE|UPDATE|DELETE&modelo=Produto&usuario=<id>
             &inicio=YYYY-MM-DD&fim=YYYY-MM-DD
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    qs = LogAuditoria.objects.select_related("usuario").order_by("-timestamp")

    acao = request.GET.get("acao")
    acoes_validas = [k for k, _ in LogAuditoria.ACAO_CHOICES]
    if acao in acoes_validas:
        qs = qs.filter(acao=acao)

    modelo = request.GET.get("modelo")
    if modelo:
        qs = qs.filter(modelo__icontains=modelo)

    usuario_id = request.GET.get("usuario")
    if usuario_id and usuario_id.isdigit():
        qs = qs.filter(usuario_id=int(usuario_id))

    inicio = request.GET.get("inicio")
    fim = request.GET.get("fim")
    if inicio:
        try:
            qs = qs.filter(timestamp__date__gte=inicio)
        except Exception:
            pass
    if fim:
        try:
            qs = qs.filter(timestamp__date__lte=fim)
        except Exception:
            pass
    else:
        dias = _get_periodo(request)
        desde = timezone.now() - timedelta(days=dias)
        qs = qs.filter(timestamp__gte=desde)

    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria"

    headers = [
        "Data/Hora", "Usuário", "Ação", "Modelo", "Objeto",
        "Objeto ID", "IP", "User Agent", "URL",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="065F46")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for log in qs:
        ws.append([
            fmt_date(log.timestamp),
            log.usuario.username if log.usuario_id else "(sistema)",
            log.acao,
            log.modelo or "",
            log.objeto_repr or "",
            log.objeto_id or "",
            log.ip or "",
            log.user_agent[:80] if log.user_agent else "",
            log.url or "",
        ])

    widths = [20, 18, 10, 20, 30, 10, 16, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    apply_borders(ws, f"A1:I{ws.max_row}")
    ws.freeze_panes = "A2"

    filename = f"auditoria_{localdate().isoformat()}.xlsx"
    return _xlsx_response(wb, filename)


# ============================================================================
# Export: Validade (PDF)
# ============================================================================
@login_required
@require_GET
def exportar_validade_pdf(request):
    """Exporta relatório de validade (lotes vencidos + vencendo) em PDF."""
    from xhtml2pdf import pisa

    hoje = localdate()
    limite = hoje + timedelta(days=90)

    lotes = (
        Lote.objects
        .filter(ativo=True, quantidade_atual__gt=0, data_validade__isnull=False)
        .select_related("produto", "produto__categoria")
        .order_by("data_validade")
    )

    faixas = {
        "VENCIDOS": [],
        "CRITICO (≤3 dias)": [],
        "ALERTA (≤7 dias)": [],
        "ATENCAO (≤30 dias)": [],
        "OK (>30 dias)": [],
    }
    for lote in lotes:
        if lote.vencido:
            faixas["VENCIDOS"].append(lote)
        else:
            dias = lote.dias_para_vencer
            if dias is None:
                continue
            if dias <= 3:
                faixas["CRITICO (≤3 dias)"].append(lote)
            elif dias <= 7:
                faixas["ALERTA (≤7 dias)"].append(lote)
            elif dias <= 30:
                faixas["ATENCAO (≤30 dias)"].append(lote)
            else:
                faixas["OK (>30 dias)"].append(lote)

    total_lotes = sum(len(v) for v in faixas.values())
    total_critico = sum(
        len(faixas[k]) for k in ("VENCIDOS", "CRITICO (≤3 dias)", "ALERTA (≤7 dias)")
    )

    html = _render_validade_html(faixas, hoje, total_lotes, total_critico)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="relatorio_validade_{hoje.isoformat()}.pdf"'
    )
    pisa.CreatePDF(html, dest=response, encoding="utf-8")
    return response


def _render_validade_html(faixas, hoje, total_lotes, total_critico):
    """Renderiza o HTML do relatório de validade (consumido por xhtml2pdf)."""
    from django.conf import settings

    empresa = getattr(settings, "EMPRESA_NOME", "Sistema de Estoque")

    def render_faixa(titulo, lotes, cor):
        if not lotes:
            return ""
        rows = ""
        for lote in lotes:
            dias = lote.dias_para_vencer
            rows += f"""
            <tr>
              <td>{lote.produto.nome}</td>
              <td>{lote.produto.codigo_interno}</td>
              <td>{lote.numero_lote}</td>
              <td class='num'>{lote.quantidade_atual} {lote.produto.unidade_medida}</td>
              <td>{lote.data_validade.strftime('%d/%m/%Y')}</td>
              <td class='num'>{dias if dias is not None else '—'}</td>
            </tr>"""
        return f"""
        <h2 style='color:{cor}; margin-top:18px;'>{titulo} ({len(lotes)})</h2>
        <table>
          <thead>
            <tr>
              <th>Produto</th><th>Código</th><th>Lote</th>
              <th>Quantidade</th><th>Validade</th><th>Dias</th>
            </tr>
          </thead>
          <tbody>{rows}</tbody>
        </table>"""

    sections = (
        render_faixa("Vencidos", faixas["VENCIDOS"], "#B91C1C")
        + render_faixa("Crítico (≤3 dias)", faixas["CRITICO (≤3 dias)"], "#DC2626")
        + render_faixa("Alerta (≤7 dias)", faixas["ALERTA (≤7 dias)"], "#D97706")
        + render_faixa("Atenção (≤30 dias)", faixas["ATENCAO (≤30 dias)"], "#2563EB")
        + render_faixa("OK (>30 dias)", faixas["OK (>30 dias)"], "#059669")
    )

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  @page {{ size: A4; margin: 2cm 1.5cm; }}
  body {{ font-family: Helvetica, Arial, sans-serif; font-size: 10pt; color: #111; }}
  h1 {{ margin: 0 0 4px 0; font-size: 18pt; color: #065F46; }}
  h2 {{ font-size: 12pt; border-bottom: 1px solid #ddd; padding-bottom: 4px; }}
  .meta {{ color: #666; font-size: 9pt; margin-bottom: 12px; }}
  .kpi {{ display: inline-block; padding: 6px 10px; background: #F3F4F6; margin-right: 8px; border-radius: 4px; }}
  .kpi strong {{ color: #065F46; font-size: 12pt; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 14px; }}
  th, td {{ border: 1px solid #999; padding: 4px 6px; text-align: left; font-size: 9pt; }}
  th {{ background: #F3F4F6; }}
  td.num {{ text-align: right; font-family: monospace; }}
  .footer {{ margin-top: 20px; color: #999; font-size: 8pt; text-align: center; }}
</style>
</head><body>
<h1>Relatório de Validade</h1>
<div class="meta">
  {empresa} · Gerado em {hoje.strftime('%d/%m/%Y')}<br>
  Janela: vencidos + até 90 dias à frente
</div>
<div>
  <span class="kpi">Lotes no relatório: <strong>{total_lotes}</strong></span>
  <span class="kpi">Requerem ação: <strong>{total_critico}</strong></span>
</div>
{sections}
<div class="footer">Estoque Cozinha · Relatório automático</div>
</body></html>"""
